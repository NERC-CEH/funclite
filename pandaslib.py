# pylint: skip-file
# this is skipped because
"""routines to manipulate array like objects like lists, tuples etc"""
from warnings import warn
import os.path as _path
from copy import deepcopy as _deepcopy

import itertools as _itertools

import pandas as pd
import statsmodels.stats.api as _sms
import pandas as _pd
# from pandas.compat import StringIO as _StringIO
import numpy as np

try:
    import xlwings as _xlwings
except Exception as err:
    print('Failed to import xlwings.\n\n%s' % err)

from sklearn.metrics import mean_squared_error as _mean_squared_error

from funclite.iolib import PrintProgress as _PrintProgress
import funclite.iolib as _iolib
from funclite.to_precision import std_notation as _std_notation
import funclite.stringslib as _stringslib
from dblib import mssql
import funclite.statslib as _statslib


# region Pandas
class MoveColumn:
    """ Move columns in pandas dataframe
    
    This is a verbatim copy of this repo ... full credit to the author.
    
    Credit: https://github.com/saadbinmunir/MoveColumn/tree/main/movecolumn
    """

    @staticmethod
    def MoveTo1(df, col: str) -> _pd.DataFrame:
        col1 = df.pop(col)
        df.insert(loc=0, column=col, value=col1)
        return df

    @staticmethod
    def MoveTo2(df, col: str) -> _pd.DataFrame:
        col1 = df.pop(col)
        df.insert(loc=1, column=col, value=col1)
        return df

    @staticmethod
    def MoveTo3(df, col: str) -> _pd.DataFrame:
        col1 = df.pop(col)
        df.insert(loc=2, column=col, value=col1)
        return df

    @staticmethod
    def MoveTo4(df, col: str) -> _pd.DataFrame:
        col1 = df.pop(col)
        df.insert(loc=3, column=col, value=col1)
        return df

    @staticmethod
    def MoveTo5(df, col: str) -> _pd.DataFrame:
        col1 = df.pop(col)
        df.insert(loc=4, column=col, value=col1)
        return df

    @staticmethod
    def MoveToLast(df, col: str) -> _pd.DataFrame:
        col1 = df.pop(col)
        df.insert(loc=len(df.columns), column=col, value=col1)
        return df

    @staticmethod
    def MoveToN(df, col, n: int) -> _pd.DataFrame:
        col1 = df.pop(col)
        df.insert(loc=n - 1, column=col, value=col1)
        return df


class IntervalHelper:
    """
    Functions to do with Intervals

    Methods:
        interval_get: Get the text for the interval for the value v, or get the Interval itself

    Notes:
        max_ is included in the intervals (unlike the range built-in)

    """

    def __init__(self, min_, max_, step):
        self._min = min_
        self._max = max_
        self._step = step
        self._IntervalIndex = _pd.IntervalIndex.from_breaks(range(min_, max_ + 1, step))

    def interval_get(self, v: (int, float), as_text: bool = True, sep: str = _stringslib.Characters.Language.emdash, as_mid: bool = False, mid_fmt_func=int, is_not_in_interval='na'):
        """
        Get the interval as text

        Args:
            v: value to test
            as_text: get the interval as text, otherwise get as a pandas.Interval
            sep: seperator for intervals.
            as_mid: Get as midpoint (ignores sep)
            mid_fmt_func: Apply this func to format the mid value str, ignored if as_mid is False
            is_not_in_interval: return this value if v is not in the class interval

        Returns:
            str: Interval as a string, suitable for writing to tables
            pandas.Interval: Get as a pandas.Interval object
            None: If v not in the intervals definied in the init

        Notes:
            If as_text is False, then we return the Cut object, irrespective of is_not_in_internal

        Examples:

            Value not in the defined Intervals

            >>> IntervalHelper().interval_get(-1)
            'na'

            Value is in the Intervals

            >>> IntervalHelper().interval_get(12)
            '0 - 25'
        """
        Cut = _pd.cut([v], self._IntervalIndex)[0]
        if as_text:
            if not _pd.notna(Cut): return is_not_in_interval
            if as_mid: return str(mid_fmt_func(Cut.mid))
            return _stringslib.filter_alphanumeric1(str(Cut), strict=True).replace(' ', sep)
        return Cut


class GroupBy:
    """Wrapper for the Pandas group_by function.

    Includes a library of aggregate functions which allow arguments
    to be passed to the function, allowing calculation of spefic metrics
    e.g. a custom percentile.

    GroupBy.PRECISION controls number formats for some functions

    Args:
        df: pandas dataframe
        groupflds: iterable list of fields to aggregate by
        valuefls: list of value fields with the data to summarise
        drop_na_in_groupby: Passed to dataframe.groupby - control if na's in the groupby fields are dropped (default) or retained
        *funcs: list of aggregate functions
        **kwargs: Accepts flatten (bool) and col_names_out (tuple,iter) for flattening and renaming the outputed dataframe respectively. col_names_out simply matches by index.

    Notes:
        Funcs can be custom static functions (e.g. static members of this class, or other scipy or numpy functions). The numpy library is exposed via GroupyBy.numpy for easy function access


    Examples:

        General example

        >>> df = <LOAD A DATA FRAME>  # noqa
        >>> GroupBy.PRECISION = 1
        >>> GB = GroupBy(df, ['fish', 'sex'], ['length', 'weight'], np.mean, np.median, GroupBy.fCI(95), GroupBy.fPercentile(25), flatten=True)  # noqa
        >>> GB.to_excel('C:\temp\tmp.xlsx', fail_if_exists=False)  # noqa

        Using col_names_out and flatten

        >>> GroupBy(df, ['country', 'region'], ['population'], GroupBy.numpy.sum, flatten=True, col_names_out=('country', 'region', 'population')).result  # noqa
        country     region      population
        UK          Midlands    10000000
    """

    PRECISION = 2
    ALPHA = 95  # percentages, 95 is a 0.05 alpha - it is this was so that the final df output col headings are ok

    numpy = np  # for convienance to easily specify aggregate funcs

    def __init__(self, df, groupflds, valueflds, *funcs, drop_na_in_groupby: bool = True, **kwargs):
        self.df = df

        if isinstance(groupflds, str): groupflds = [groupflds]
        if isinstance(valueflds, str): valueflds = [valueflds]
        self._groupflds = groupflds
        self._valueflds = valueflds

        self._drop_na_in_groupby = drop_na_in_groupby
        self._flatten = kwargs.get('flatten')
        self._col_names_out = kwargs.get('col_names_out')

        self._funcs = funcs
        self.result = None
        self.execute()

    def execute(self):
        """
        Execute the aggregate and return the result as a dataframe.
        Use this method to return a flattened dataset

        Returns: _pd.Dataframe: The results in a pandas dataframe pandas

        Notes:
            flatten is forced if col_names_out evaluates to True (otherwise we get an index error)
        """
        allfuncs = self._funcs * len(self._valueflds)  # noqa
        d = {}
        for v in self._valueflds:
            d[v] = self._funcs

        self.result = self.df.groupby(self._groupflds, dropna=self._drop_na_in_groupby).agg(d)
        if self._flatten or self._col_names_out:  # col_names_out needs a flattened result
            # self.result.reset_index(inplace=True)
            self.result.columns = self.result.columns.to_flat_index().str.join('_')
            self.result.reset_index(inplace=True)
        if self._col_names_out:
            for i, v in enumerate(self._col_names_out):
                self.result.columns.values[i] = v
        return self.result

    def to_excel(self, fname, sheetname='Sheet1', fail_if_exists=True, openfld=True):
        """save as excel"""
        if not isinstance(self.result, _pd.DataFrame):
            warn('Aggregate results do not exist')
            return
        fname = _path.normpath(fname)
        fld, f, _ = _iolib.get_file_parts2(fname)
        if _iolib.file_exists(fname) and fail_if_exists:
            raise FileExistsError('File %s exists and fail_if_exists==True' % f)

        assert isinstance(self.result, _pd.DataFrame)
        self.result.to_excel(fname, sheetname)
        if openfld:
            try:
                _iolib.folder_open(fld)
            except:
                pass

    @staticmethod
    def fPercentile(n):
        def percentile_(data):
            return np.percentile(data, n)

        percentile_.__name__ = 'percentile_%s' % n
        return percentile_

    @staticmethod
    def fMSE(pred):
        """Mean squared error
        Assess quality of estimator
        MSE = 1/n * sum(x - xmodel)^2

        pred is the expected value, i.e. model mean
        """

        def mse_(data):
            ndpred = np.zeros(data.shape[0]) + pred
            return _mean_squared_error(ndpred, data)

        mse_.__name__ = 'mse_%s' % pred
        return mse_

    @staticmethod
    def fRMSE(pred):
        """Root mean squared error.
        Assess quality of estimator.
        RMSE is measured on same scale and units as the dependent variable.
        RMSE =  sqrt( 1/n * sum(x - xmodel)^2 )
        pred is the expected value, i.e. model mean
        """

        def rmse_(data):
            ndpred = np.zeros(data.shape[0]) + pred
            x = _mean_squared_error(ndpred, data) ** 0.5
            return x

        rmse_.__name__ = 'rmse_%s' % pred
        return rmse_

    @staticmethod
    def fCI(interval=95):
        def ci_(data):
            a = _sms.DescrStatsW(data).tconfint_mean((100 - interval) * 0.01)
            return (max(a) - min(a)) / 2.

        ci_.__name__ = 'ci_%s' % interval
        return ci_

    @staticmethod
    def fCILower(interval=95):
        def CILower_(data):
            l, _ = _sms.DescrStatsW(data).tconfint_mean((100 - interval) * 0.01)
            return l

        CILower_.__name__ = 'CILower_%s' % interval
        return CILower_

    @staticmethod
    def CIUpper(interval=95):
        def CIUpper_(data):
            l, _ = _sms.DescrStatsW(data).tconfint_mean((100 - interval) * 0.01)
            return l

        CIUpper_.__name__ = 'CIUpper_%s' % interval
        return CIUpper_

    @staticmethod
    def fCI_str(interval=95):
        """formatted str version"""

        def CI_str_(data):
            l, _ = _sms.DescrStatsW(data).tconfint_mean((100 - interval) * 0.01)
            m = np.mean(data)
            s = 'M=%s %s%% CIs [%s, %s]' % (_std_notation(m, GroupBy.PRECISION), interval, _std_notation(l, GroupBy.PRECISION), _std_notation(u, GroupBy.PRECISION))  # noqa
            # return m, m-h, m+h
            return s

        CI_str_.__name__ = 'CI_str_%s' % interval
        return CI_str_

    @staticmethod
    def n(data):
        d = np.array(data)
        return np.count_nonzero(d)

    @staticmethod
    def fMeanSD_str(data):
        d = np.array(data)
        m, sd = np.mean(d), np.std(d)
        s = '%s %s%s' % (_std_notation(m, GroupBy.PRECISION), _plus_minus(), _std_notation(sd, GroupBy.PRECISION))
        return s

    @staticmethod
    def fCIUpperFinite(group_N, two_tailed=True):
        last = 0
        # https://www.statisticshowto.datasciencecentral.com/finite-population-correction-factor/
        """group_N is an iter of length equal the number of groups.
        
        Each value in group_N is equal to the population number from which the aggregate
        sample was drawn.

        The order of group_N must match the sort order
        of the groups (which is the order groups appear as variable data below.
        
        The order of data will be a standard ordered sort on orignal dataframe's aggregate groups.

        See test_pandaslib for a worked example.
        """

        def CIUpperFinite_(data):
            interval = (100 - GroupBy.ALPHA) * 0.01  # pass interval as 95 so cols have nice names
            nonlocal last
            n = group_N[last]  # the population nr
            last += 1
            d = np.array(data)
            m, se, ciabs, ci_lower, ci_upper = _statslib.finite_population_stats(d, n, interval, two_tailed)
            return ci_upper

        CIUpperFinite_.__name__ = 'CIUpperFinite_%s' % GroupBy.ALPHA
        return CIUpperFinite_

    @staticmethod
    def fCILowerFinite(group_N, two_tailed=True):
        last = 0
        # https://www.statisticshowto.datasciencecentral.com/finite-population-correction-factor/
        """group_N is an iter of length equal the number of groups.
        
        Each value in group_N is equal to the population number from which the aggregate
        sample was drawn.

        The order of group_N must match the sort order
        of the groups (which is the order groups appear as variable data below.
        
        The order of data will be a standard ordered sort on orignal dataframe's aggregate groups.

        See test_pandaslib for a worked example.
        """

        def CILowerFinite_(data):
            nonlocal last
            interval = (100 - GroupBy.ALPHA) * 0.01  # pass interval as 95 so cols have nice names
            n = group_N[last]  # the population nr
            last += 1
            d = np.array(data)
            m, se, ciabs, ci_lower, ci_upper = _statslib.finite_population_stats(d, n, interval, two_tailed)
            return ci_lower

        CILowerFinite_.__name__ = 'CILowerFinite_%s' % GroupBy.ALPHA
        return CILowerFinite_

    @staticmethod
    def fSEFinite(group_N, two_tailed=True):
        last = 0
        # https://www.statisticshowto.datasciencecentral.com/finite-population-correction-factor/
        """group_N is an iter of length equal the number of groups.
        
        Each value in group_N is equal to the population number from which the aggregate
        sample was drawn.

        The order of group_N must match the sort order
        of the groups (which is the order groups appear as variable data below.
        
        The order of data will be a standard ordered sort on orignal dataframe's aggregate groups.

        See test_pandaslib for a worked example.
        """

        def SELowerFinite_(data):
            nonlocal last
            interval = (100 - GroupBy.ALPHA) * 0.01  # pass interval as 95 so cols have nice names
            n = group_N[last]  # the population nr
            last += 1
            d = np.array(data)
            m, se, ciabs, ci_lower, ci_upper = _statslib.finite_population_stats(d, n, interval, two_tailed)
            return ci_lower

        SELowerFinite_.__name__ = 'SELowerFinite_%s' % GroupBy.ALPHA
        return SELowerFinite_


def df_to_ndarray(df):
    """(dataframe)->ndarray
    Return a dataframe as a numpy array
    """
    return df.as_matrix([x for x in df.columns])


def vstack(*args):
    """vertically stack dataframses
    Just a call to pandas.concat, but i kept forgetting what it was called
    """
    return _pd.concat([df for df in args])


def col_append(df, col_name):
    """(df,str)->df
    df is BYREF
    adds a column to dataframe filling it
    with np.NaN values.
    """
    df.loc[:, col_name] = _pd.Series(_pd.np.nan, index=df.index)


def col_calculate_percent_by_group(df: _pd.DataFrame,
                                   perc_col_name: str,
                                   group_by_cols: (str, list, None),
                                   value_col: str,
                                   as_proportion: bool = False,
                                   round_func=lambda v: np.round(v, decimals=1),
                                   inplace=False) -> (None, _pd.DataFrame):
    """
    Add a percentage or proportion calculated column, split/stratified by the values in another column

    Args:
        df: dataframe
        perc_col_name: new col to add, assigned percent values
        group_by_cols: The groupings within which to calculate percent values, pass None, to apply across whole dataframe
        value_col: The name of the col with the values
        as_proportion: Return as proportion rather then percent
        round_func: A function accepting a single float value which rounds, this is used to round the result, useful when output is used for reporting purposes. Pass none to leave value as-is.
        inplace: Inplace, otherwise return

    Returns:
        None: If in place
        pandas.DataFrame: If inplace is false

    Notes:
        Also see the general method col_calculate_by_group

    Examples:

        Percent of "count" by county as percentage, returning a new dataframe

        Grouping on single column

        >>> dfres = pd.DataFrame({'county': ['Berks', 'Yorks'] * 3, 'count': [10,10,30, 42,7,1]})  # noqa
        >>> col_calculate_percent_by_group(dfres, 'perc', 'county', 'count')
        county  perc
        Berks   50
        Yorks   50
        ...


        Grouping on multiple columns (source dataframe not shown)

        >>> col_calculate_percent_by_group(dfres, 'perc', ['county', 'country'], 'count')
        county  country perc
        Berks   England 50
        Yorks   England 50
        Mon     Wales   10
        ...
    """
    if group_by_cols:
        denom = df.groupby(group_by_cols)[value_col].transform('sum') / (1 if as_proportion else 100)  # check this in debug
    else:
        denom = df[value_col].sum() / (1 if as_proportion else 100)

    if inplace:
        df[perc_col_name] = df[value_col] / denom
        if round_func: df[perc_col_name] = df[perc_col_name].apply(round_func)
        return

    dfcpy = df.copy()
    dfcpy[perc_col_name] = dfcpy[value_col] / denom
    if round_func: dfcpy[perc_col_name] = dfcpy[perc_col_name].apply(round_func)
    return dfcpy  # noqa


def col_calculate_by_group(df: _pd.DataFrame,
                           result_col_name: str,
                           group_by_col: str, value_col: str,
                           transform, round_func=lambda v: np.round(v, decimals=1),
                           inplace=False) -> (None, _pd.DataFrame):
    """
    Add a percentage or proportion calculated column, split/stratified by the values in another column

    Args:
        df: dataframe
        result_col_name: new col to add, assigned percent values
        group_by_col: The groupings within which to calculate percent values
        value_col: The name of the col with the values
        transform: An argument accepted by Series.transform. Can be a function, str, list or dict. Some supported strings are 'sum', 'max', 'min'. See https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.Series.transform.html.
        round_func: A function accepting a single float value which rounds, this is used to round the result, useful when output is used for reporting purposes. Pass none to leave value as-is.
        inplace: Inplace, otherwise return

    Returns:
        None: If in place
        pandas.DataFrame: If inplace is false

    Notes:
        Also see the specific method col_calculate_percent_by_group

    Examples:

        max of "count" by county, returning a new dataframe

        >>> dfres = pd.DataFrame({'county': ['Berks', 'Yorks'] * 3, 'count': [10,10,30, 42,7,1]})  # noqa
        >>> col_calculate_percent_by_group(dfres, 'max_count', 'county', 'count', transform='max')
        county  max_count
        Berks   30
        Yorks   42
    """
    if not inplace:
        df[result_col_name] = df[value_col] / df.groupby(group_by_col)[value_col].transform(transform)
        if round_func: df[result_col_name] = df[result_col_name].apply(round_func)
        return

    dfcpy = df.copy()
    dfcpy[result_col_name] = dfcpy[value_col] / dfcpy.groupby(group_by_col)[value_col].transform(transform)
    if round_func: df[result_col_name] = df[result_col_name].apply(round_func)
    return dfcpy  # noqa


def col_append_nan_fill(df, col_name):
    """(df,str)->df
    df is BYREF
    adds a column to dataframe filling it
    with np.NaN values.
    """
    col_append(df, col_name)


def col_append_fill(df, col_name, f):
    """(df,str,any)->df
    df is BYREF
    adds a column to dataframe filling it with value f
    If f is None, filled with NaN
    """
    if f is None:
        df.loc[:, col_name] = _pd.Series(_pd.np.nan, index=df.index)
    else:
        df.loc[:, col_name] = _pd.Series(f, index=df.index)


def col_append_rand_fill(df, col_name, lower=0, upper=1):
    """(df,str,any)->df
    df is BYREF
    adds a column to dataframe filling it with random values from a standard normal
    """
    df[col_name] = np.random.choice(range(lower, upper), df.shape[0])


def col_calculate_new(df, func, new_col_name, *args, progress_init_msg='\n'):
    """(pd.df, function, str, the named arguments for function)
    1) Adds a new column called col_name
    2) calls func with args by position,  where args are the row indices for the values
    3) Row indexes are ZERO based
    4) Consider using apply or similiar for simple use cases

    df = pd.dataframe({'a':[1,2],'b'=[10,20]})

    DF=
    a   b
    1   10
    2   20

    def f(a, b):
        return a*b

    func = f
    col_calculate_new(df, func, 'product', 0, 1)

    DF=
    a   b   product
    1   10  10
    2   20  40
    """
    assert isinstance(df, _pd.DataFrame)
    if new_col_name in df.columns:
        raise Exception('Column %s already exists in the dataframe.' % new_col_name)
    col_append(df, new_col_name)

    args = list(args)
    args = _list_flatten(args)
    PP = _PrintProgress(len(df.index), init_msg=progress_init_msg)
    for i, row in df.iterrows():
        PP.increment()
        rowvals = []
        for x in args:
            if row[x] is None:
                vv = None
            elif np.isnan(row[x]):
                vv = None
            else:
                vv = row[x]
            rowvals.append(vv)
        v = func(*rowvals)
        df.set_value(i, new_col_name, v)


def df_to_dict_keycol_multivalues(df: _pd.DataFrame, key_col: str, val_col: str) -> dict:
    """
    Create dict from dataframe.
    Keys are created from key_col, and values are added to a list
    against each key.

    Args:
        df (pandas.DataFrame): The dataframe
        key_col (str): The col to use as keys.
        val_col (str): The col to use as values

    Returns:
        dict: Diction where keys are unique values in df[key_col], and key values are a list of all values where lookup matches the key col (see example)

    Examples:
        >>> dfr = _pd.DataFrame({'permission': ['Given', 'Not Given', 'Not Given', 'Not Given'], 'plotid':[1,2,3,4]})
        >>> dfr
          permission  plotid
        0      Given       1
        1  Not Given       2
        2  Not Given       3
        3  Not Given       4\n\n

        >>> df_to_dict_keycol_multivalues(dfr, 'permission', 'plotid')
        {'Given': [1], 'Not Given': [2, 3, 4]}
    """
    # TODO: test/debug df_to_dict_keycol_multivalues
    keys = df[key_col].to_list()
    vals = df[val_col].to_list()
    dct = {k: [] for k in set(keys)}
    for i, k in enumerate(keys):
        dct[k].append(vals[i])
    return dct


def col_exists(df, col_name):
    """(str)->bool
    """
    return col_name in df.columns


def col_index(df, col_name):
    """(df, str)->int
    Given col return index
    Returns None if doesnt exist
    """
    if col_exists(df, col_name):
        return df.columns.get_loc(col_name)
    return None


def cols_get_indexes_from_names(df, *args):
    """df, str args->tuple
    Given list if strings get the corresponding
    column indexes and return as a tuple
    """
    return [col_index(df, x) for x in args]


def readfld(v, default=None):
    """return default if v is a pandas null
    """
    return default if _pd.isnull(v) else v


# endregion


def df_mssql(sql, dbname, server='(local)', port=1433, security='integrated', user='', pw=''):
    """(str, str, str, str, str)-> pandas.dataframe
    Get a pandas dataframe from SQL Server

    sql: the sql to execute to get the data
    dbname: database name
    server:server identifier
    security: integrated or sqlserver
    user,pw: SQL server user authentication details
    """
    with mssql.Conn(dbname, server, port=port, security=security, user=user, pw=pw) as cnn:
        df = _pd.read_sql(sql, cnn)
    return df


def df_fromstring(str_, sep=',', header=0, names=None, **args):  # noqa
    """(str, str, bool, dict) -> pandas.dataframe
    Convert a python string into a dataframe.

    Pass names as an array with header=None when
    there are no header names. As set, the first
    row is assumed to contain col names
    """
    # df = _pd.read_csv(_StringIO(str_), sep=sep, header=header, names=names, engine='python', **args)
    # return df
    raise NotImplementedError


def df_lower(df: _pd.DataFrame, inplace=True):
    """
    Convert cols to lower case

    Args:
        df: df
        inplace: inplace or return a copy

    Returns:
        the df
    """
    if inplace:
        df.columns = df.columns.str.lower()
        return
    ddf = df.copy()
    ddf.columns = ddf.columns.str.lower()
    return ddf


def df_flatten_cols(df: _pd.DataFrame, inplace: bool = True) -> (None, _pd.DataFrame):
    """
    Try and flatten multiindex cols to single list of cols.

    Args:
        df (_pd.DataFrame): Pandas DF to flatten
        inplace (bool): Alter df, otherwise return an flattened dataframe

    Returns:
        _pd.DataFrame: IF inplace is false
        None: If inplace is True, df is altered
    """
    if inplace:
        df.columns = df.columns.to_flat_index().str.join('_')
        df.reset_index(inplace=True)
    else:
        ddf = df.copy()
        ddf.columns = ddf.columns.to_flat_index().str.join('_')
        ddf.reset_index(inplace=True)
        return ddf


def df_from_dict(d):
    """(dict) -> pandas.dataframe
    Build a datafrom from a dict. Keys are col headings, values are entries.
    Supports unequal length values, and values in (set, list, tuple)

    Args:
        d: dictionary

    Examples:
    >>>df_from_dict({'a':[1], 'b':[1,2]})
        a   b
    0   1   1
    1   NaN 2
    """
    return _pd.DataFrame(dict([(k, _pd.Series(list(v))) for k, v in d.items()]))


def pandas_join_multi(dfs: list, key: (str, list), **kwargs) -> _pd.DataFrame:
    """
    Joins all dataframes by a _common column

    Args:
        dfs: list or tuple of dataframes
        key: candidate key column (supports compound "keys")
        kwargs:

    Returns:
        pandas.DataFrame: The joined dataframes

    Notes:
        Assumes all key columns are named the same

    Examples:

        Single key column

        >>> pandas_join_multi([df1, df2, df3], key='objectid')  # noqa

        Compound key column

        >>> pandas_join_multi([df1, df2, df3], key=['cola', 'colb'])  # noqa
    """
    df_first = dfs[0]
    for df in dfs[1:]:
        df_first = pandas_join(df_first, df, from_key=key, to_key=key, **kwargs)
    return df_first


def pandas_join(from_: _pd.DataFrame, to_: _pd.DataFrame, from_key: str, to_key: str, drop_wildcard: str = '__', how='left', **kwargs) -> _pd.DataFrame:
    """
    Join two pandas dataframes base on two named key cols.

    Args:
        from_: Datafrome
        to_: left join to this dataframe
        from_key: key in "from" table
        to_key: key in "to" table
        drop_wildcard: matched cols will be
        how: "left", "inner", "right"
        kwargs: Keyword args to pass to pandas join func

    Returns:
        pandas dataframe from the join

    Notes:
        See https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.join.html for kwargs
    """
    join = from_.set_index(from_key).join(to_.set_index(to_key), how=how, rsuffix=drop_wildcard, **kwargs)
    if drop_wildcard:
        join.drop(list(join.filter(regex=drop_wildcard)), axis=1, inplace=True)  # drop duplicate cols
    join = join.reset_index()
    return join


def df_to_dict_as_records_flatten(lst: list) -> dict:
    """
    Takes a two column dataframe, exported to_dict
    as records and outputs as a dict of
    key-value pairs.

    Args:
        lst (list): A list of dicts, where each dict is a single "row",\ne.g. [{'crn': 'A0012017', 'permission': 'Not Given'}, {'crn': 'A0011574', 'permission': 'Given'}, ...]

    Returns:
        dict: Dictionary with the values of the first col as the key with second as the value (see example)

    Notes:
        Also see df_to_dict_as_records_flatten, which takes a dataframe and converts it to correctly formatted dict.
        Useful for arcapi.data.field_update_from_dict.

    Examples:
        >>> d = df_crn_permissions_status_responded_xlsx().to_dict(orient='records')  # noqa
        >>> print(d)
        [{'crn': 'A0012017', 'permission': 'Not Given'}, {'crn': 'A0011574', 'permission': 'Given'}, ...]\n
        >>> df_to_dict_as_records_flatten(d)
        {'A0012017': 'Not Given', 'A0011574': 'Given' ....}
    """
    dd = dict()
    for d in lst:
        lst_items = list(d.items())
        dd[lst_items[0][1]] = lst_items[1][1]
    return dd


def df_to_dict_as_records_flatten1(df: _pd.DataFrame, cols: (list[str], None) = None) -> dict:
    """
    Takes a two column dataframe, converts to exported to_dict
    as records and outputs as a dict of
    key-value pairs.

    Args:
        df (pandas.DataFrame): A dataframe in standard format
        cols (list[str], None): A 2n list of columns to get key/value pairs from. If None, takes first two columns, with first col as the key col

    Raises:
        ValueError: If len(cols) != 2

    Returns:
        dict: Dictionary with the values of the first col as the key with second as the value (see example)

    Notes:
        See also df_to_dict_as_records_flatten, which this is a small wrapper around.
        Useful for arcapi.data.field_update_from_dict

    Examples:
        >>> df = _pd.DataFrame({'col1':['a','b','c'], 'col2':[1,2,3], 'col3':[10,20,30]})  # noqa

        Passing cols
        >>> df_to_dict_as_records_flatten1(df, ['col1', 'col3'])
        {'a':10, 'b':20, 'c':30}

        Do not pass cols, defaults to first two cols in dataframe
        >>> df_to_dict_as_records_flatten1(df)
        {'a':1, 'b':2, 'c':3}
    """
    if cols and len(cols) != 2:
        raise ValueError('The cols list %s should contain two column names' % cols)

    if cols:
        return df_to_dict_as_records_flatten(df[[cols[0], cols[1]]].to_dict(orient='records'))
    return df_to_dict_as_records_flatten(df[[df.columns[0], df.columns[1]]].to_dict(orient='records'))


def df_tally_df(cols: (list[str], tuple[str]), *args):
    """
    This gets a full tally dataframe.

    That is, it performs a product (itertools.product) operation on the input arg iterables to get all value combinations.

    From that, a dataframe is made with len(args) columns, with every combination of the values in args
    across the table.

    This is a familiar concept in databases. Tally tables are used to join related datasets on the tally table to ensure
    that all possible values have a record that is not otherwise ommited when joining.

    Args:
        cols: iterable of column names
        *args: list or tuples of columunar data

    Returns:
        _pd.DataFrame

    Examples:

        >>> df_tally_df(('a', 'b', 'c'), [1,2], ['b','bb'], [1.1, 2.2])
        a       b       c
        'a'     1       1.1
        'a'     1       2.2
        'a'     2       1.1
        'a'     2       2.2
        'b'     ...     ...
    """
    return _pd.DataFrame(_itertools.product(*args), columns=cols).reset_index()


def dfs_to_excel(dfs: list[pd.DataFrame], save_to: str, sheet_names: (None, list[str]) = None, overwrite: bool = True, show_progress: bool = False, **kwargs) -> None:
    """
    Export multiple dataframes to a single workbook

    Args:
        dfs (list[pd.DataFrame]): Dataframes to export to save_to
        save_to (str): Filename to export dfs to
        sheet_names (None, list[str]):
        overwrite (bool): Allow overwrite
        show_progress (bool): Show progress
        kwargs: passed to each to_excel call on each df

    Raises:
        FileExistsError: If the file exists and overwrite was False
        ValueError: If sheet_names was not None and number of sheet_names was different from number of dataframes

    Returns:
        None

    Examples:

        Export a dataframe to 3 worksheets, and specify index and startrow kwargs to pass to DataFrame.to_excel

        >>> df = pd.DataFrame({1:[1,2,3], 2:[2,3,4]}
        >>> dfs_to_excel([df, df, df], 'C:/temp/my.xlsx', ['df1', 'df2', 'df3'], index=False, startrow=1)
    """
    save_to = _path.normpath(save_to)
    if not sheet_names:
        sheet_names = ['Sheet%s' % s for s in range(1, len(dfs) + 1)]
    else:
        if len(dfs) != len(sheet_names):
            raise ValueError('If you pass sheet_names, then the number of names must be the same as the number of dataframes')

    if not overwrite and _iolib.file_exists(save_to):
        raise FileExistsError('File %s already exists and overwrite was False' % save_to)
    _iolib.file_delete(save_to)

    if show_progress:
        PP = _iolib.PrintProgress(maximum=len(dfs), init_msg='Exporting dataframes to excel file %s' % save_to)

    with _pd.ExcelWriter(save_to) as EW:
        for i, sheet in enumerate(sheet_names):
            dfs[i].to_excel(EW, sheet_name=sheet, **kwargs)
            if show_progress:
                PP.increment()  # noqa


def df_to_excel_append(filename, df, sheet_name='Sheet1', startrow=None, startcol=None,
                       truncate_sheet=False, resizeColumns=True, na_rep='NA', **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Args:
        filename: File path or existing ExcelWriter (Example: '/path/to/file.xlsx')
        df: dataframe to save to workbook
        sheet_name : Name of sheet which will contain DataFrame.

        startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...

        startcol: col to start

        truncate_sheet : truncate (remove and recreate) [sheet_name] before writing DataFrame to Excel file

        resizeColumns: default = True . It resize all columns based on cell content width
        to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()
        na_rep: default = 'NA'. If, instead of NaN, you want blank cells, just edit as follows: na_rep=''

    Returns: None



    Credits:
        Originally from: https://stackoverflow.com/a/66599924/5585800


        Current helper function generated by [Baggio]: https://stackoverflow.com/users/14302009/baggio?tab=profile
        Contributions to the current helper function: https://stackoverflow.com/users/4046632/buran?tab=profile
        Original helper function: (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)


        Features of the new helper function:
        1) Now it works with python 3.9 and latest versions of pandas and openpxl
        ---> Fixed the error: "zipfile.BadZipFile: File is not a zip file".
        2) Now It resize all columns based on cell content width AND all variables will be visible (SEE "resizeColumns")
        3) You can handle NaN,  if you want that NaN are displayed as NaN or as empty cells (SEE "na_rep")
        4) Added "startcol", you can decide to start to write from specific column, oterwise will start from col = 0
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook
    filename = _path.normpath(filename)
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    try:
        _ = open(filename)
        # Do something with the file
    except IOError:
        # print("File not accessible")
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename)

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError  # noqa
    except NameError:
        FileNotFoundError = IOError  # noqa

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        # startrow = -1
        startrow = 0

    if startcol is None:
        startcol = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, na_rep=na_rep, **to_excel_kwargs)

    if resizeColumns:

        ws = writer.book[sheet_name]

        def auto_format_cell_width(ws_):
            for letter in range(1, ws_.max_column):
                maximum_value = 0
                for cell in ws_[get_column_letter(letter)]:
                    val_to_check = len(str(cell.value))
                    if val_to_check > maximum_value:
                        maximum_value = val_to_check
                ws_.column_dimensions[get_column_letter(letter)].width = maximum_value + 2

        auto_format_cell_width(ws)

    # save the workbook
    writer.save()


def excel_table_as_df(workbook: str, worksheet: (str, int), table: (str, int)) -> _pd.DataFrame:
    """
    Get an excel listobject (table) as a dataframe

    Args:
        workbook (str): Workbook path
        worksheet (str, int): worksheet name or index
        table (str, int): table name or index

    Raises:
        KeyError: If the workbook, worksheet or table do not exist

    Returns:
        pandas.DataFrame: The table as a pandas dataframe
    """
    workbook = _path.normpath(workbook)

    with _xlwings.App(visible=False) as App:
        _ = App.books.open(workbook)
        rng = App.books[_path.basename(workbook)].sheets[worksheet].tables[table].range
        df: _pd.DataFrame = rng.expand().options(_pd.DataFrame).value
        df.reset_index(inplace=True)
    return df


def excel_diff(wb1: str, sht1: str, wb2: str = None, sht2: str = None, tbl1: str = None, tbl2: str = None, key_cols: tuple = (), diff_cols: tuple = (), diff_only: bool = False) -> _pd.DataFrame:
    """
    Difference between two excel worksheets.

    Args:
        wb1 (str):
        sht1 (str):
        wb2 (str, None):
        sht2 (str):
        tbl1 (str):
        tbl2 (str):
        key_cols (tuple[str], None):
        diff_cols (tuple[str], None):
        diff_only (bool): Only return rows where any single cell had a difference

    Raises:
        ValueError: If tbl1 or tbl2 are passed, but the other tbl arg evaluates to False. Or if wb1 or wb2 are not defined but tbl1 and tbl2 are (we need the workbook to get the table).

    Returns:
        pandas.DataFrame: Differences
    """
    # TODO: Complete excel_diff.
    raise NotImplementedError("Not yet implemented")
    if not wb2: wb2 = wb1
    if not sht2: sht2 = sht1

    if tbl1 and not tbl2 or not tbl1 and tbl2:
        raise ValueError('If comparing tables, arguments must be provided for tbl1 and tbl2')
    elif tbl1 and tbl2:
        if not sht1 or not sht2:
            raise ValueError('If comparing tables, wb1 and wb2 must be defined.')

    if tbl1 and tbl2:
        pass
    else:
        df1 = _pd.read_excel('excel1.xlsx')  # noqa
        df2 = pd.read_excel('excel2.xlsx')  # noqa
    # order by key cols
    # retain key cols as seperate dataframe
    # check that values in key cols are unique
    # need to find missing keys in each and add to two dataframes (fill in row gaps)
    # difference between ordered dataframe cols defined by diff_cols, or NOT keycols if not diff_cols
    # join on results
    # Filter rows where all diff_cols are NaN according to diff_only
    # consider adding a field is_changed, flagged as 0 if no change in the row, else 1

    diff_df1_vals = df1[df1 != df2]
    diff_df2_vals = df2[df1 != df2]

    return diff_df1_vals, diff_df2_vals  # noqa


# ---------------
# Helper funcs
# ---------------
def _list_flatten(items, seqtypes=(list, tuple)):
    """flatten a list

    **beware, this is also by ref**
    """
    citems = _deepcopy(items)
    for i, dummy in enumerate(citems):
        while i < len(citems) and isinstance(citems[i], seqtypes):
            citems[i:i + 1] = citems[i]
    return citems


def _plus_minus():
    """get plus minus"""
    return str(u"\u00B1")
